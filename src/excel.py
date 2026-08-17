import os
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import config

ACTIVE_SHEET   = "Active Outreach"
REPLIED_SHEET  = "Replied"
ARCHIVED_SHEET = "Archived"

COL = {
    "company":           1,
    "contact_name":      2,
    "email":             3,
    "subject":           4,
    "initial_body":      5,
    "followup_body":     6,
    "date_sent":         7,
    "followup_count":    8,
    "last_followup":     9,
    "status":            10,
    "message_id":        11,
    "conversation_id":   12,
}

ACTIVE_COLS   = 12
REPLIED_COLS  = 4
ARCHIVED_COLS = 4

THIN   = Side(style="thin", color="BDD7EE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CELL_FONT = Font(name="Arial", size=10)
LEFT  = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALT_A = PatternFill("solid", start_color="DEEAF1")
ALT_B = PatternFill("solid", start_color="FFFFFF")


def _style_row(ws, row_idx, n_cols):
    fill = ALT_A if row_idx % 2 == 0 else ALT_B
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill      = fill
        cell.font      = CELL_FONT
        cell.border    = BORDER
        cell.alignment = LEFT


def _excel_path():
    return config.path("excel_path", "outreach.xlsx")


def _follow_up_interval() -> int:
    return config.get_int("settings", "follow_up_interval_days", 3)


def _max_follow_ups() -> int:
    return config.get_int("settings", "max_follow_ups", 5)


def _load():
    path = _excel_path()
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Tracker not found at {path}\n"
            "Copy outreach.template.xlsx to outreach.xlsx, or fix settings.excel_path."
        )
    return load_workbook(path)


def _save(wb):
    path = _excel_path()
    lock = os.path.join(os.path.dirname(path), "~$" + os.path.basename(path))
    if os.path.exists(lock):
        raise RuntimeError(
            f"{os.path.basename(path)} is open in Excel — close it before running, "
            "otherwise your edits and the bot's will overwrite each other."
        )
    wb.save(path)


def _cell_value(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if isinstance(val, datetime):
        return val.date()
    return val


def _to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _next_empty_row(ws) -> int:
    for row in range(2, ws.max_row + 2):
        if not any(ws.cell(row, col).value for col in range(1, 5)):
            return row
    return ws.max_row + 1


def _row_to_dict(ws, row: int) -> dict:
    return {
        "row":             row,
        "company":         _cell_value(ws, row, COL["company"]),
        "contact_name":    _cell_value(ws, row, COL["contact_name"]),
        "email":           _cell_value(ws, row, COL["email"]),
        "subject":         _cell_value(ws, row, COL["subject"]),
        "initial_body":    _cell_value(ws, row, COL["initial_body"]),
        "followup_body":   _cell_value(ws, row, COL["followup_body"]),
        "date_sent":       _to_date(_cell_value(ws, row, COL["date_sent"])),
        "followup_count":  _cell_value(ws, row, COL["followup_count"]) or 0,
        "last_followup":   _to_date(_cell_value(ws, row, COL["last_followup"])),
        "status":          _cell_value(ws, row, COL["status"]),
        "message_id":      _cell_value(ws, row, COL["message_id"]),
        "conversation_id": _cell_value(ws, row, COL["conversation_id"]),
    }


# ── Read contacts ─────────────────────────────────────────────────────────────

def get_pending_contacts() -> list[dict]:
    wb = _load()
    ws = wb[ACTIVE_SHEET]
    contacts = []
    for row in range(2, ws.max_row + 1):
        status = _cell_value(ws, row, COL["status"])
        if str(status).strip().lower() == "pending":
            contacts.append(_row_to_dict(ws, row))
    return contacts


def get_followup_candidates() -> list[dict]:
    wb = _load()
    ws = wb[ACTIVE_SHEET]
    today = date.today()
    interval = _follow_up_interval()
    max_ups = _max_follow_ups()
    candidates = []
    for row in range(2, ws.max_row + 1):
        status = str(_cell_value(ws, row, COL["status"]) or "").strip().lower()
        if status not in ("sent", "follow-up sent"):
            continue
        count = _cell_value(ws, row, COL["followup_count"]) or 0
        if count >= max_ups:
            continue
        last = _to_date(_cell_value(ws, row, COL["last_followup"]))
        sent = _to_date(_cell_value(ws, row, COL["date_sent"]))
        reference_date = last if last else sent
        if reference_date and (today - reference_date).days >= interval:
            candidates.append(_row_to_dict(ws, row))
    return candidates


def get_all_active_emails() -> dict[str, int]:
    wb = _load()
    ws = wb[ACTIVE_SHEET]
    result = {}
    for row in range(2, ws.max_row + 1):
        email = _cell_value(ws, row, COL["email"])
        if email:
            result[str(email).strip().lower()] = row
    return result


def get_empty_conversation_ids() -> list[dict]:
    """Return rows where conversation_id is empty but status is Sent or Follow-up Sent."""
    wb = _load()
    ws = wb[ACTIVE_SHEET]
    rows = []
    for row in range(2, ws.max_row + 1):
        status  = str(_cell_value(ws, row, COL["status"]) or "").strip().lower()
        conv_id = _cell_value(ws, row, COL["conversation_id"])
        if status in ("sent", "follow-up sent") and not conv_id:
            rows.append(_row_to_dict(ws, row))
    return rows


# ── Write updates ─────────────────────────────────────────────────────────────

def mark_sent(row: int, message_id: str, conversation_id: str = "", sent_date: date = None):
    """sent_date defaults to today; scheduled sends pass their delivery date so
    follow-ups are counted from when the mail actually goes out, not from now."""
    wb = _load()
    ws = wb[ACTIVE_SHEET]
    ws.cell(row=row, column=COL["date_sent"]).value      = sent_date or date.today()
    ws.cell(row=row, column=COL["followup_count"]).value = 0
    ws.cell(row=row, column=COL["status"]).value         = "Sent"
    ws.cell(row=row, column=COL["message_id"]).value     = message_id
    if conversation_id:
        ws.cell(row=row, column=COL["conversation_id"]).value = conversation_id
    _save(wb)


def mark_followup_sent(row: int):
    wb = _load()
    ws = wb[ACTIVE_SHEET]
    count = ws.cell(row=row, column=COL["followup_count"]).value or 0
    ws.cell(row=row, column=COL["followup_count"]).value = count + 1
    ws.cell(row=row, column=COL["last_followup"]).value  = date.today()
    ws.cell(row=row, column=COL["status"]).value         = "Follow-up Sent"
    _save(wb)


def backfill_conversation_id(row: int, conv_id: str):
    wb = _load()
    ws = wb[ACTIVE_SHEET]
    ws.cell(row=row, column=COL["conversation_id"]).value = conv_id
    _save(wb)


def move_to_replied(row: int):
    wb = _load()
    ws_active  = wb[ACTIVE_SHEET]
    ws_replied = wb[REPLIED_SHEET]
    d = _row_to_dict(ws_active, row)
    next_row = _next_empty_row(ws_replied)
    ws_replied.cell(next_row, 1).value = d["company"]
    ws_replied.cell(next_row, 2).value = d["contact_name"]
    ws_replied.cell(next_row, 3).value = d["email"]
    ws_replied.cell(next_row, 4).value = date.today()
    _style_row(ws_replied, next_row, REPLIED_COLS)
    ws_active.delete_rows(row)
    _save(wb)


def move_to_archived(row: int):
    wb = _load()
    ws_active   = wb[ACTIVE_SHEET]
    ws_archived = wb[ARCHIVED_SHEET]
    d = _row_to_dict(ws_active, row)
    next_row = _next_empty_row(ws_archived)
    ws_archived.cell(next_row, 1).value = d["company"]
    ws_archived.cell(next_row, 2).value = d["contact_name"]
    ws_archived.cell(next_row, 3).value = d["email"]
    ws_archived.cell(next_row, 4).value = date.today()
    _style_row(ws_archived, next_row, ARCHIVED_COLS)
    ws_active.delete_rows(row)
    _save(wb)


# ── Fix email address ─────────────────────────────────────────────────────────

def find_contact(search_term: str) -> list[dict]:
    wb = _load()
    ws = wb[ACTIVE_SHEET]
    term = search_term.strip().lower()
    matches = []
    for row in range(2, ws.max_row + 1):
        company = str(_cell_value(ws, row, COL["company"]) or "").lower()
        name    = str(_cell_value(ws, row, COL["contact_name"]) or "").lower()
        if term in company or term in name:
            matches.append(_row_to_dict(ws, row))
    return matches


def fix_email(row: int, new_email: str):
    wb = _load()
    ws = wb[ACTIVE_SHEET]
    ws.cell(row=row, column=COL["email"]).value            = new_email.strip()
    ws.cell(row=row, column=COL["date_sent"]).value        = None
    ws.cell(row=row, column=COL["followup_count"]).value   = 0
    ws.cell(row=row, column=COL["last_followup"]).value    = None
    ws.cell(row=row, column=COL["status"]).value           = "Pending"
    ws.cell(row=row, column=COL["message_id"]).value       = None
    ws.cell(row=row, column=COL["conversation_id"]).value  = None
    _save(wb)


# ── Status summary ────────────────────────────────────────────────────────────

def get_status_summary() -> dict:
    wb = _load()
    ws_active   = wb[ACTIVE_SHEET]
    ws_replied  = wb[REPLIED_SHEET]
    ws_archived = wb[ARCHIVED_SHEET]
    today = date.today()
    interval = _follow_up_interval()
    max_ups = _max_follow_ups()
    active_count = pending = sent = followup_sent = due_today = 0

    for row in range(2, ws_active.max_row + 1):
        status = str(_cell_value(ws_active, row, COL["status"]) or "").strip()
        if not status:
            continue
        active_count += 1
        if status.lower() == "pending":
            pending += 1
        elif status.lower() == "sent":
            sent += 1
            count = _cell_value(ws_active, row, COL["followup_count"]) or 0
            if count < max_ups:
                sent_date = _to_date(_cell_value(ws_active, row, COL["date_sent"]))
                if sent_date and (today - sent_date).days >= interval:
                    due_today += 1
        elif status.lower() == "follow-up sent":
            followup_sent += 1
            count = _cell_value(ws_active, row, COL["followup_count"]) or 0
            if count < max_ups:
                last = _to_date(_cell_value(ws_active, row, COL["last_followup"]))
                if last and (today - last).days >= interval:
                    due_today += 1

    def _count_rows(ws):
        return sum(
            1 for row in range(2, ws.max_row + 1)
            if any(ws.cell(row, col).value for col in range(1, 5))
        )

    return {
        "active":        active_count,
        "pending":       pending,
        "sent":          sent,
        "followup_sent": followup_sent,
        "replied":       _count_rows(ws_replied),
        "archived":      _count_rows(ws_archived),
        "due_today":     due_today,
    }


# ── Import from Claude-generated Excel ───────────────────────────────────────

def import_from_file(source_path: str) -> dict:
    from openpyxl import load_workbook as lw

    src_wb = lw(source_path, read_only=True)
    src_ws = src_wb.active

    headers = []
    header_row = 1
    for i, row in enumerate(src_ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if any(str(c).strip() == "Company" for c in row if c):
            headers = [str(cell).strip() if cell else "" for cell in row]
            header_row = i
            break

    if not headers:
        return {"imported": 0, "skipped": 0, "reasons": ["Could not find header row in source file"]}

    HEADER_MAP = {
        "Company":            "company",
        "Contact Name":       "contact_name",
        "Email":              "email",
        "Subject":            "subject",
        "Initial Email Body": "initial_body",
        "Follow-up Body":     "followup_body",
        "Status":             "status",
    }

    col_index = {}
    for i, h in enumerate(headers):
        if h in HEADER_MAP:
            col_index[HEADER_MAP[h]] = i

    dst_wb = _load()
    dst_ws = dst_wb[ACTIVE_SHEET]

    imported = 0
    skipped  = 0
    reasons  = []

    for row_vals in src_ws.iter_rows(min_row=header_row + 1, values_only=True):
        def get(key):
            idx = col_index.get(key)
            return str(row_vals[idx]).strip() if idx is not None and row_vals[idx] is not None else ""

        # Trailing blank rows are an artefact of how the sheet was authored,
        # not real contacts — ignore them instead of reporting phantom skips.
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue

        email = get("email")
        if not email:
            skipped += 1
            reasons.append(f"Skipped '{get('company') or get('contact_name')}' — no email")
            continue

        next_row = _next_empty_row(dst_ws)
        dst_ws.cell(next_row, COL["company"]).value        = get("company")
        dst_ws.cell(next_row, COL["contact_name"]).value   = get("contact_name")
        dst_ws.cell(next_row, COL["email"]).value          = email
        dst_ws.cell(next_row, COL["subject"]).value        = get("subject")
        dst_ws.cell(next_row, COL["initial_body"]).value   = get("initial_body")
        dst_ws.cell(next_row, COL["followup_body"]).value  = get("followup_body")
        dst_ws.cell(next_row, COL["followup_count"]).value = 0
        dst_ws.cell(next_row, COL["status"]).value         = get("status") or "Pending"
        _style_row(dst_ws, next_row, ACTIVE_COLS)
        imported += 1

    src_wb.close()
    _save(dst_wb)
    return {"imported": imported, "skipped": skipped, "reasons": reasons}